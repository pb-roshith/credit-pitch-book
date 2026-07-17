from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from psycopg import sql

from database import ensure_database, get_connection
from table_config import TABLE_FILES


def split_table_name(table_name):
    parts = table_name.split('.', 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return 'credit_dossier', table_name


def normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def read_excel(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError(f'Workbook is empty: {path}')

    columns = [str(value).strip() for value in rows[0] if value is not None and str(value).strip()]
    if not columns:
        raise ValueError(f'Workbook has no header row: {path}')

    data_rows = []
    for row in rows[1:]:
        values = list(row[: len(columns)])
        if all(value is None for value in values):
            continue
        data_rows.append([normalize_cell(value) for value in values])

    return columns, data_rows


def create_table(conn, table_name, columns):
    schema_name, bare_table_name = split_table_name(table_name)
    conn.execute(sql.SQL('CREATE SCHEMA IF NOT EXISTS {};').format(sql.Identifier(schema_name)))
    column_definitions = [
        sql.SQL('{} TEXT').format(sql.Identifier(column))
        for column in columns
    ]
    query = sql.SQL('CREATE TABLE {} ({})').format(
        sql.Identifier(schema_name, bare_table_name),
        sql.SQL(', ').join(column_definitions),
    )
    with conn.cursor() as cur:
        cur.execute(sql.SQL('DROP TABLE IF EXISTS {};').format(sql.Identifier(schema_name, bare_table_name)))
        cur.execute(query)


def insert_rows(conn, table_name, columns, rows):
    if not rows:
        return

    schema_name, bare_table_name = split_table_name(table_name)
    placeholders = sql.SQL(', ').join(sql.Placeholder() for _ in columns)
    query = sql.SQL('INSERT INTO {} ({}) VALUES ({})').format(
        sql.Identifier(schema_name, bare_table_name),
        sql.SQL(', ').join(sql.Identifier(column) for column in columns),
        placeholders,
    )
    with conn.cursor() as cur:
        cur.executemany(query, rows)


def load_all_tables():
    ensure_database()
    loaded = {}
    with get_connection() as conn:
        for table_name, path in TABLE_FILES.items():
            columns, rows = read_excel(path)
            create_table(conn, table_name, columns)
            insert_rows(conn, table_name, columns, rows)
            loaded[table_name] = {'columns': columns, 'rows': len(rows)}
        conn.commit()
    return loaded


if __name__ == '__main__':
    result = load_all_tables()
    for table_name, metadata in result.items():
        print(f'{table_name}: {metadata["rows"]} rows, {len(metadata["columns"])} columns')
